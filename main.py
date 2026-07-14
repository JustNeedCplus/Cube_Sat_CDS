#!/usr/bin/env python3
"""CubeSat Builder - Subsystem Sizing, BOM & Budget Tool for CubeSats/Small Satellites"""

import webview
import json
import os
import sys
import math
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


# ============================================================
# CONSTANTS & CATALOGUES
# ============================================================

# CubeSat form factor limits
CUBESAT_STANDARDS = {
    '1U': {'mass_kg': 1.33, 'size_mm': (100, 100, 113.5), 'volume_L': 1.0},
    '1.5U': {'mass_kg': 2.0, 'size_mm': (100, 100, 170.2), 'volume_L': 1.5},
    '2U': {'mass_kg': 2.66, 'size_mm': (100, 100, 227.0), 'volume_L': 2.0},
    '3U': {'mass_kg': 4.0, 'size_mm': (100, 100, 340.5), 'volume_L': 3.0},
    '6U': {'mass_kg': 12.0, 'size_mm': (200, 100, 340.5), 'volume_L': 6.0},
    '12U': {'mass_kg': 24.0, 'size_mm': (200, 200, 340.5), 'volume_L': 12.0},
    '16U': {'mass_kg': 32.0, 'size_mm': (200, 200, 454.0), 'volume_L': 16.0},
    '27U': {'mass_kg': 54.0, 'size_mm': (300, 300, 340.5), 'volume_L': 27.0},
}

# Orbit parameters
ORBIT_PRESETS = {
    'LEO-400': {'altitude_km': 400, 'period_min': 92.4, 'eclipse_min': 36.0, 'sunlight_min': 56.4},
    'LEO-500': {'altitude_km': 500, 'period_min': 94.6, 'eclipse_min': 35.3, 'sunlight_min': 59.3},
    'LEO-600': {'altitude_km': 600, 'period_min': 96.7, 'eclipse_min': 34.5, 'sunlight_min': 62.2},
    'SSO-500': {'altitude_km': 500, 'period_min': 94.6, 'eclipse_min': 35.3, 'sunlight_min': 59.3},
    'SSO-600': {'altitude_km': 600, 'period_min': 96.7, 'eclipse_min': 34.5, 'sunlight_min': 62.2},
    'ISS': {'altitude_km': 408, 'period_min': 92.7, 'eclipse_min': 36.0, 'sunlight_min': 56.7},
}

# Solar cell types
SOLAR_CELLS = {
    'SpectroLab UTJ': {'efficiency': 0.295, 'degradation_per_year': 0.02, 'w_per_m2': 400, 'cost_per_w': 300},
    'SpectroLab XTJ': {'efficiency': 0.305, 'degradation_per_year': 0.015, 'w_per_m2': 415, 'cost_per_w': 400},
    'Azur 3G30A': {'efficiency': 0.30, 'degradation_per_year': 0.02, 'w_per_m2': 407, 'cost_per_w': 350},
    'SolAero ZTJ': {'efficiency': 0.295, 'degradation_per_year': 0.018, 'w_per_m2': 400, 'cost_per_w': 320},
    'CESI CTJ30': {'efficiency': 0.30, 'degradation_per_year': 0.02, 'w_per_m2': 407, 'cost_per_w': 280},
    'Generic Si': {'efficiency': 0.18, 'degradation_per_year': 0.03, 'w_per_m2': 245, 'cost_per_w': 80},
}

# COTS EPS (Electrical Power System) catalogue
# cost_usd values are illustrative public-range estimates only — not quotes.
EPS_CATALOGUE = [
    {'vendor': 'GOMspace', 'model': 'NanoPower P31u', 'form': '1U-3U', 'channels': 6, 'bus_v': '3.3/5.0', 'max_solar_w': 7.0, 'battery_wh': 38.5, 'mass_g': 75, 'cost_usd': 6500},
    {'vendor': 'GOMspace', 'model': 'NanoPower P60', 'form': '3U-12U', 'channels': 12, 'bus_v': '3.3/5.0/12.0', 'max_solar_w': 40.0, 'battery_wh': 80.0, 'mass_g': 235, 'cost_usd': 15000},
    {'vendor': 'GOMspace', 'model': 'NanoPower BPX', 'form': 'Battery', 'channels': 0, 'bus_v': 'N/A', 'max_solar_w': 0, 'battery_wh': 38.5, 'mass_g': 210, 'cost_usd': 4500},
    {'vendor': 'Endurosat', 'model': 'EPS I', 'form': '1U-3U', 'channels': 8, 'bus_v': '3.3/5.0', 'max_solar_w': 10.0, 'battery_wh': 20.0, 'mass_g': 86, 'cost_usd': 5000},
    {'vendor': 'Endurosat', 'model': 'EPS II', 'form': '3U-6U', 'channels': 10, 'bus_v': '3.3/5.0/12.0', 'max_solar_w': 30.0, 'battery_wh': 40.0, 'mass_g': 150, 'cost_usd': 9000},
    {'vendor': 'Clyde Space', 'model': 'Optimus-30', 'form': '1U-3U', 'channels': 5, 'bus_v': '3.3/5.0', 'max_solar_w': 30.0, 'battery_wh': 30.0, 'mass_g': 96, 'cost_usd': 7000},
    {'vendor': 'Clyde Space', 'model': 'Optimus-60', 'form': '3U-12U', 'channels': 10, 'bus_v': '3.3/5.0/12.0', 'max_solar_w': 60.0, 'battery_wh': 60.0, 'mass_g': 200, 'cost_usd': 14000},
    {'vendor': 'ISIS', 'model': 'iEPS', 'form': '1U-3U', 'channels': 4, 'bus_v': '3.3/5.0', 'max_solar_w': 8.0, 'battery_wh': 20.7, 'mass_g': 94, 'cost_usd': 5500},
    {'vendor': 'NanoAvionics', 'model': 'EPSL', 'form': '1U-6U', 'channels': 8, 'bus_v': '3.3/5.0/12.0', 'max_solar_w': 20.0, 'battery_wh': 40.0, 'mass_g': 110, 'cost_usd': 7500},
    {'vendor': 'AAC Clyde', 'model': 'Starbuck-Micro', 'form': '1U-6U', 'channels': 6, 'bus_v': '3.3/5.0/12.0', 'max_solar_w': 25.0, 'battery_wh': 45.0, 'mass_g': 130, 'cost_usd': 8000},
]

# COTS Radio/Transceiver catalogue
RADIO_CATALOGUE = [
    {'vendor': 'Endurosat', 'model': 'UHF Transceiver II', 'band': 'UHF', 'freq_mhz': 437, 'tx_power_w': 2.0, 'data_rate_kbps': 19.2, 'mass_g': 70, 'power_w': 4.0, 'cost_usd': 4500},
    {'vendor': 'Endurosat', 'model': 'S-Band Transmitter', 'band': 'S-Band', 'freq_mhz': 2400, 'tx_power_w': 2.0, 'data_rate_kbps': 2000, 'mass_g': 85, 'power_w': 8.0, 'cost_usd': 9000},
    {'vendor': 'Endurosat', 'model': 'X-Band Transmitter', 'band': 'X-Band', 'freq_mhz': 8200, 'tx_power_w': 2.0, 'data_rate_kbps': 50000, 'mass_g': 300, 'power_w': 18.0, 'cost_usd': 25000},
    {'vendor': 'GOMspace', 'model': 'NanoCom AX100', 'band': 'UHF', 'freq_mhz': 437, 'tx_power_w': 1.0, 'data_rate_kbps': 19.2, 'mass_g': 30, 'power_w': 3.5, 'cost_usd': 5000},
    {'vendor': 'GOMspace', 'model': 'NanoCom SR2000', 'band': 'S-Band', 'freq_mhz': 2200, 'tx_power_w': 1.0, 'data_rate_kbps': 1000, 'mass_g': 55, 'power_w': 6.0, 'cost_usd': 12000},
    {'vendor': 'ISIS', 'model': 'TRXVU', 'band': 'UHF/VHF', 'freq_mhz': 437, 'tx_power_w': 0.5, 'data_rate_kbps': 9.6, 'mass_g': 85, 'power_w': 4.0, 'cost_usd': 6000},
    {'vendor': 'ISIS', 'model': 'TXS', 'band': 'S-Band', 'freq_mhz': 2400, 'tx_power_w': 1.0, 'data_rate_kbps': 5000, 'mass_g': 70, 'power_w': 10.0, 'cost_usd': 15000},
    {'vendor': 'NanoAvionics', 'model': 'SatCOM UHF', 'band': 'UHF', 'freq_mhz': 437, 'tx_power_w': 2.0, 'data_rate_kbps': 19.2, 'mass_g': 60, 'power_w': 5.0, 'cost_usd': 5500},
    {'vendor': 'Addvalue', 'model': 'IDRS', 'band': 'L-Band', 'freq_mhz': 1600, 'tx_power_w': 1.5, 'data_rate_kbps': 9.6, 'mass_g': 200, 'power_w': 8.0, 'cost_usd': 20000},
    {'vendor': 'Tethers Unlimited', 'model': 'SWIFT-SLX', 'band': 'X-Band', 'freq_mhz': 8200, 'tx_power_w': 4.0, 'data_rate_kbps': 100000, 'mass_g': 450, 'power_w': 25.0, 'cost_usd': 35000},
]

# COTS ADCS (Attitude Determination and Control) catalogue
ADCS_CATALOGUE = [
    {'vendor': 'CubeSpace', 'model': 'CubeADCS 3-Axis', 'type': '3-axis', 'pointing_deg': 1.0, 'mass_g': 300, 'power_w': 1.5, 'form': '1U-6U', 'cost_usd': 35000},
    {'vendor': 'NanoAvionics', 'model': 'SatBus ADCS', 'type': '3-axis', 'pointing_deg': 0.5, 'mass_g': 450, 'power_w': 2.0, 'form': '3U-12U', 'cost_usd': 45000},
    {'vendor': 'NewSpace Systems', 'model': 'NFSS-411', 'type': 'Sun sensor', 'pointing_deg': 0.5, 'mass_g': 5, 'power_w': 0.03, 'form': 'Any', 'cost_usd': 3000},
    {'vendor': 'Hyperion', 'model': 'ST400', 'type': 'Star tracker', 'pointing_deg': 0.002, 'mass_g': 230, 'power_w': 1.0, 'form': '3U+', 'cost_usd': 55000},
    {'vendor': 'GOMspace', 'model': 'NanoMind A3200', 'type': 'OBC+ADCS', 'pointing_deg': 5.0, 'mass_g': 25, 'power_w': 0.4, 'form': '1U-3U', 'cost_usd': 8000},
    {'vendor': 'ISIS', 'model': 'iMTQ', 'type': 'Magnetorquer', 'pointing_deg': 10.0, 'mass_g': 196, 'power_w': 1.2, 'form': '1U-3U', 'cost_usd': 5000},
    {'vendor': 'Blue Canyon', 'model': 'XACT-15', 'type': '3-axis', 'pointing_deg': 0.003, 'mass_g': 885, 'power_w': 2.0, 'form': '3U-12U', 'cost_usd': 85000},
    {'vendor': 'Tensor Tech', 'model': 'RW4', 'type': 'Reaction wheel', 'pointing_deg': 0.1, 'mass_g': 150, 'power_w': 0.5, 'form': '1U-6U', 'cost_usd': 12000},
]

# OBC (On-Board Computer) catalogue
OBC_CATALOGUE = [
    {'vendor': 'GOMspace', 'model': 'NanoMind A3200', 'processor': 'AVR32', 'ram_mb': 0.5, 'storage_mb': 2, 'mass_g': 25, 'power_w': 0.4, 'cost_usd': 8000},
    {'vendor': 'ISIS', 'model': 'iOBC', 'processor': 'ARM9', 'ram_mb': 64, 'storage_mb': 256, 'mass_g': 94, 'power_w': 0.4, 'cost_usd': 9000},
    {'vendor': 'NanoAvionics', 'model': 'SatBus 3C2', 'processor': 'ARM Cortex-A', 'ram_mb': 512, 'storage_mb': 4000, 'mass_g': 45, 'power_w': 0.8, 'cost_usd': 12000},
    {'vendor': 'Endurosat', 'model': 'OBC', 'processor': 'ARM Cortex-M7', 'ram_mb': 1, 'storage_mb': 16, 'mass_g': 50, 'power_w': 0.3, 'cost_usd': 5000},
    {'vendor': 'Unibap', 'model': 'iX5-100', 'processor': 'Intel x86', 'ram_mb': 8000, 'storage_mb': 128000, 'mass_g': 350, 'power_w': 15.0, 'cost_usd': 45000},
    {'vendor': 'AAC Clyde', 'model': 'KRYTEN-M3', 'processor': 'ARM Cortex-A53', 'ram_mb': 2000, 'storage_mb': 32000, 'mass_g': 90, 'power_w': 3.0, 'cost_usd': 18000},
    {'vendor': 'Xiphos', 'model': 'Q7S', 'processor': 'Zynq SoC', 'ram_mb': 1000, 'storage_mb': 32000, 'mass_g': 70, 'power_w': 3.5, 'cost_usd': 22000},
]

# Standard subsystem list for power budget
DEFAULT_SUBSYSTEMS = [
    {'name': 'OBC', 'safe_w': 0.4, 'nominal_w': 0.8, 'peak_w': 1.5, 'duty_pct': 100, 'category': 'Avionics'},
    {'name': 'EPS', 'safe_w': 0.2, 'nominal_w': 0.5, 'peak_w': 0.5, 'duty_pct': 100, 'category': 'Power'},
    {'name': 'UHF Radio', 'safe_w': 0.5, 'nominal_w': 1.0, 'peak_w': 4.0, 'duty_pct': 15, 'category': 'Comms'},
    {'name': 'ADCS', 'safe_w': 0.0, 'nominal_w': 1.5, 'peak_w': 3.0, 'duty_pct': 100, 'category': 'ADCS'},
    {'name': 'Payload', 'safe_w': 0.0, 'nominal_w': 5.0, 'peak_w': 10.0, 'duty_pct': 30, 'category': 'Payload'},
    {'name': 'Thermal', 'safe_w': 1.0, 'nominal_w': 2.0, 'peak_w': 5.0, 'duty_pct': 50, 'category': 'Thermal'},
]


# ============================================================
# CALCULATION FUNCTIONS
# ============================================================

def calc_orbit_params(altitude_km):
    """Calculate orbital parameters from altitude."""
    R_earth = 6371.0  # km
    mu = 398600.4  # km^3/s^2
    r = R_earth + altitude_km
    period_s = 2 * math.pi * math.sqrt(r**3 / mu)
    period_min = period_s / 60.0
    # Eclipse fraction (circular orbit, worst case)
    rho = math.asin(R_earth / r)
    eclipse_frac = rho / math.pi
    eclipse_min = period_min * eclipse_frac
    sunlight_min = period_min - eclipse_min
    return {
        'altitude_km': altitude_km,
        'period_min': round(period_min, 1),
        'eclipse_min': round(eclipse_min, 1),
        'sunlight_min': round(sunlight_min, 1),
        'eclipse_frac': round(eclipse_frac, 3),
        'orbits_per_day': round(1440 / period_min, 1)
    }


def calc_power_budget(subsystems, orbit):
    """Calculate power budget across all modes."""
    safe_total = sum(s.get('safe_w', 0) for s in subsystems)
    nominal_total = sum(s.get('nominal_w', 0) for s in subsystems)
    peak_total = sum(s.get('peak_w', 0) for s in subsystems)
    # Average power considering duty cycles
    avg_total = sum(s.get('nominal_w', 0) * s.get('duty_pct', 100) / 100.0 for s in subsystems)
    # Power needed from solar (must charge battery during sunlight for eclipse)
    if orbit:
        eclipse_frac = orbit.get('eclipse_frac', 0.39)
        sunlight_frac = 1.0 - eclipse_frac
        # Solar must provide orbit-average power plus charge battery for eclipse
        # Assuming 90% battery charge efficiency, 95% EPS efficiency
        charge_eff = 0.90
        eps_eff = 0.95
        solar_required = avg_total / eps_eff
        if sunlight_frac > 0:
            solar_required += (avg_total * eclipse_frac) / (sunlight_frac * charge_eff * eps_eff)
        battery_required_wh = avg_total * orbit.get('eclipse_min', 35) / 60.0 / charge_eff
    else:
        solar_required = avg_total * 1.5
        battery_required_wh = avg_total * 0.6  # 36 min eclipse

    return {
        'safe_w': round(safe_total, 2),
        'nominal_w': round(nominal_total, 2),
        'peak_w': round(peak_total, 2),
        'average_w': round(avg_total, 2),
        'solar_required_w': round(solar_required, 2),
        'battery_required_wh': round(battery_required_wh, 2),
        'margin_note': '20% margin recommended'
    }


def calc_solar_panel(required_w, cell_type, mission_years, margin_pct=20):
    """Size solar panels."""
    cell = SOLAR_CELLS.get(cell_type, SOLAR_CELLS['Azur 3G30A'])
    # End of life degradation
    eol_factor = (1 - cell['degradation_per_year']) ** mission_years
    # Cosine loss average (assume body-mounted, ~70% effective)
    cosine_factor = 0.70
    # Required power with margin
    required_with_margin = required_w * (1 + margin_pct / 100)
    # BOL power per m2
    bol_w_per_m2 = cell['w_per_m2']
    eol_w_per_m2 = bol_w_per_m2 * eol_factor * cosine_factor
    # Area required
    area_m2 = required_with_margin / eol_w_per_m2 if eol_w_per_m2 > 0 else 0
    # Cost
    cost = required_with_margin * cell['cost_per_w']
    # Number of standard cells (approx 30x80mm = 0.0024 m2 per cell)
    cell_area = 0.0024
    num_cells = math.ceil(area_m2 / cell_area) if area_m2 > 0 else 0

    return {
        'cell_type': cell_type,
        'efficiency': cell['efficiency'],
        'eol_factor': round(eol_factor, 3),
        'required_w_bol': round(required_with_margin / (eol_factor * cosine_factor), 1),
        'required_w_eol': round(required_with_margin, 1),
        'area_m2': round(area_m2, 4),
        'area_cm2': round(area_m2 * 10000, 1),
        'num_cells': num_cells,
        'cost_usd': round(cost, 0),
        'mass_kg': round(area_m2 * 2.5, 3),  # ~2.5 kg/m2 for panel assembly
    }


def calc_link_budget(altitude_km, freq_mhz, tx_power_w, tx_gain_dbi, rx_gain_dbi,
                     data_rate_kbps, required_eb_n0=10):
    """Calculate communication link budget."""
    c = 3e8  # speed of light
    freq_hz = freq_mhz * 1e6
    wavelength = c / freq_hz
    # Slant range (worst case at 5 deg elevation)
    R_earth = 6371.0
    h = altitude_km
    elev_rad = math.radians(5)
    slant_range_km = -R_earth * math.sin(elev_rad) + math.sqrt(
        (R_earth * math.sin(elev_rad))**2 + 2 * R_earth * h + h**2)
    slant_range_m = slant_range_km * 1000
    # Free space path loss
    fspl_db = 20 * math.log10(4 * math.pi * slant_range_m / wavelength)
    # EIRP
    tx_power_dbw = 10 * math.log10(tx_power_w)
    eirp_dbw = tx_power_dbw + tx_gain_dbi
    # Received power
    rx_power_dbw = eirp_dbw - fspl_db + rx_gain_dbi
    # Noise
    noise_temp_k = 300  # system noise temperature
    k_bolt = 1.38e-23
    noise_density_dbw = 10 * math.log10(k_bolt * noise_temp_k)
    # Data rate
    data_rate_bps = data_rate_kbps * 1000
    data_rate_db = 10 * math.log10(data_rate_bps)
    # Eb/N0
    eb_n0 = rx_power_dbw - noise_density_dbw - data_rate_db
    # Margin
    margin = eb_n0 - required_eb_n0

    return {
        'slant_range_km': round(slant_range_km, 1),
        'fspl_db': round(fspl_db, 1),
        'eirp_dbw': round(eirp_dbw, 1),
        'rx_power_dbw': round(rx_power_dbw, 1),
        'eb_n0_db': round(eb_n0, 1),
        'required_eb_n0': required_eb_n0,
        'margin_db': round(margin, 1),
        'margin_ok': margin >= 3.0,
        'data_rate_kbps': data_rate_kbps,
        'pass_duration_min': round(2 * math.degrees(math.acos(R_earth * math.cos(elev_rad) / (R_earth + h))) / 360 * calc_orbit_params(altitude_km)['period_min'], 1),
    }


def match_eps(solar_w, battery_wh, form_factor):
    """Find matching EPS from catalogue."""
    matches = []
    for eps in EPS_CATALOGUE:
        if eps['max_solar_w'] >= solar_w and eps['battery_wh'] >= battery_wh:
            matches.append(eps)
    matches.sort(key=lambda x: x['cost_usd'])
    return matches


def match_radio(band=None, min_data_rate=0):
    """Find matching radios from catalogue."""
    matches = []
    for r in RADIO_CATALOGUE:
        if band and band.upper() not in r['band'].upper():
            continue
        if r['data_rate_kbps'] >= min_data_rate:
            matches.append(r)
    matches.sort(key=lambda x: x['data_rate_kbps'])
    return matches


def _safe_float(val, default=0):
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


# ============================================================
# API CLASS
# ============================================================

class Api:
    def __init__(self):
        self.window = None
        self.mission_name = 'New CubeSat Mission'
        self.form_factor = '3U'
        self.orbit = ORBIT_PRESETS['LEO-500'].copy()
        self.mission_years = 2
        self.subsystems = [dict(s) for s in DEFAULT_SUBSYSTEMS]
        self.mass_items = []
        self.bom_items = []
        self.selected_eps = None
        self.selected_radio = None
        self.selected_adcs = None
        self.selected_obc = None
        self.solar_config = {'cell_type': 'Azur 3G30A', 'margin_pct': 20}

    # --- Mission Setup ---
    def set_mission(self, name, form_factor, orbit_preset, altitude_km, mission_years):
        """Configure the mission."""
        self.mission_name = name
        self.form_factor = form_factor
        self.mission_years = int(mission_years)
        if orbit_preset in ORBIT_PRESETS:
            self.orbit = ORBIT_PRESETS[orbit_preset].copy()
        else:
            self.orbit = calc_orbit_params(float(altitude_km))
        return json.dumps({"status": "ok", "orbit": self.orbit, "form_factor": form_factor,
                           "limits": CUBESAT_STANDARDS.get(form_factor, {})})

    def get_mission(self):
        """Get current mission config."""
        return json.dumps({
            "status": "ok",
            "name": self.mission_name,
            "form_factor": self.form_factor,
            "orbit": self.orbit,
            "mission_years": self.mission_years,
            "limits": CUBESAT_STANDARDS.get(self.form_factor, {}),
            "orbit_presets": list(ORBIT_PRESETS.keys()),
            "form_factors": list(CUBESAT_STANDARDS.keys()),
            "solar_cells": list(SOLAR_CELLS.keys()),
        })

    # --- Power Budget ---
    def get_power_budget(self):
        """Calculate and return power budget."""
        budget = calc_power_budget(self.subsystems, self.orbit)
        solar = calc_solar_panel(
            budget['solar_required_w'],
            self.solar_config.get('cell_type', 'Azur 3G30A'),
            self.mission_years,
            self.solar_config.get('margin_pct', 20)
        )
        eps_matches = match_eps(solar['required_w_bol'], budget['battery_required_wh'], self.form_factor)
        return json.dumps({
            "status": "ok",
            "subsystems": self.subsystems,
            "budget": budget,
            "solar": solar,
            "orbit": self.orbit,
            "eps_matches": eps_matches[:5],
            "selected_eps": self.selected_eps,
        })

    def update_subsystem(self, index, field, value):
        """Update a subsystem power entry."""
        idx = int(index)
        if 0 <= idx < len(self.subsystems):
            if field in ('safe_w', 'nominal_w', 'peak_w', 'duty_pct'):
                value = _safe_float(value, 0)
            self.subsystems[idx][field] = value
            return json.dumps({"status": "ok"})
        return json.dumps({"status": "error", "message": "Invalid index"})

    def add_subsystem(self, name, category, safe_w, nominal_w, peak_w, duty_pct):
        """Add a subsystem to power budget."""
        self.subsystems.append({
            'name': name, 'category': category,
            'safe_w': _safe_float(safe_w), 'nominal_w': _safe_float(nominal_w),
            'peak_w': _safe_float(peak_w), 'duty_pct': _safe_float(duty_pct, 100)
        })
        return json.dumps({"status": "ok", "count": len(self.subsystems)})

    def remove_subsystem(self, index):
        """Remove a subsystem."""
        idx = int(index)
        if 0 <= idx < len(self.subsystems):
            self.subsystems.pop(idx)
            return json.dumps({"status": "ok"})
        return json.dumps({"status": "error"})

    def select_eps(self, model):
        """Select an EPS from catalogue."""
        for eps in EPS_CATALOGUE:
            if eps['model'] == model:
                self.selected_eps = eps
                self._update_bom('EPS', eps['vendor'], eps['model'], eps['mass_g'], eps['cost_usd'])
                return json.dumps({"status": "ok", "eps": eps})
        return json.dumps({"status": "error", "message": "EPS not found"})

    def set_solar_config(self, cell_type, margin_pct):
        """Set solar panel configuration."""
        self.solar_config = {'cell_type': cell_type, 'margin_pct': _safe_float(margin_pct, 20)}
        return json.dumps({"status": "ok"})

    # --- Comms Link Budget ---
    def get_link_budget(self, band, tx_gain, rx_gain):
        """Calculate link budget and return matching radios."""
        tx_g = _safe_float(tx_gain, 0)
        rx_g = _safe_float(rx_gain, 10)
        radios = match_radio(band)
        results = []
        for r in radios:
            lb = calc_link_budget(
                self.orbit['altitude_km'], r['freq_mhz'], r['tx_power_w'],
                tx_g, rx_g, r['data_rate_kbps']
            )
            results.append({**r, 'link': lb})
        return json.dumps({
            "status": "ok",
            "results": results,
            "orbit": self.orbit,
            "selected_radio": self.selected_radio,
        })

    def select_radio(self, model):
        """Select a radio from catalogue."""
        for r in RADIO_CATALOGUE:
            if r['model'] == model:
                self.selected_radio = r
                self._update_bom('Comms', r['vendor'], r['model'], r['mass_g'], r['cost_usd'])
                return json.dumps({"status": "ok", "radio": r})
        return json.dumps({"status": "error"})

    # --- Mass Budget ---
    def get_mass_budget(self):
        """Get mass budget summary."""
        limits = CUBESAT_STANDARDS.get(self.form_factor, {})
        max_mass = limits.get('mass_kg', 4.0)
        items = self._build_mass_list()
        total = sum(i['mass_kg'] for i in items)
        margin = max_mass - total
        margin_pct = (margin / max_mass * 100) if max_mass > 0 else 0
        return json.dumps({
            "status": "ok",
            "items": items,
            "total_kg": round(total, 3),
            "max_kg": max_mass,
            "margin_kg": round(margin, 3),
            "margin_pct": round(margin_pct, 1),
            "form_factor": self.form_factor,
        })

    def add_mass_item(self, name, category, mass_g, notes):
        """Add a custom mass item."""
        self.mass_items.append({
            'name': name, 'category': category,
            'mass_g': _safe_float(mass_g), 'notes': notes, 'source': 'Manual'
        })
        return json.dumps({"status": "ok"})

    def remove_mass_item(self, index):
        """Remove a custom mass item."""
        idx = int(index)
        if 0 <= idx < len(self.mass_items):
            self.mass_items.pop(idx)
            return json.dumps({"status": "ok"})
        return json.dumps({"status": "error"})

    def _build_mass_list(self):
        """Build complete mass list from selected components + manual items."""
        items = []
        if self.selected_eps:
            items.append({'name': self.selected_eps['model'], 'category': 'EPS',
                         'mass_kg': self.selected_eps['mass_g'] / 1000, 'source': 'Catalogue'})
        if self.selected_radio:
            items.append({'name': self.selected_radio['model'], 'category': 'Comms',
                         'mass_kg': self.selected_radio['mass_g'] / 1000, 'source': 'Catalogue'})
        if self.selected_adcs:
            items.append({'name': self.selected_adcs['model'], 'category': 'ADCS',
                         'mass_kg': self.selected_adcs['mass_g'] / 1000, 'source': 'Catalogue'})
        if self.selected_obc:
            items.append({'name': self.selected_obc['model'], 'category': 'OBC',
                         'mass_kg': self.selected_obc['mass_g'] / 1000, 'source': 'Catalogue'})
        # Solar panels mass
        budget = calc_power_budget(self.subsystems, self.orbit)
        solar = calc_solar_panel(budget['solar_required_w'], self.solar_config.get('cell_type', 'Azur 3G30A'),
                                 self.mission_years, self.solar_config.get('margin_pct', 20))
        if solar['mass_kg'] > 0:
            items.append({'name': f"Solar Panels ({solar['cell_type']})", 'category': 'Power',
                         'mass_kg': solar['mass_kg'], 'source': 'Calculated'})
        # Manual items
        for mi in self.mass_items:
            items.append({'name': mi['name'], 'category': mi['category'],
                         'mass_kg': mi['mass_g'] / 1000, 'source': mi.get('source', 'Manual')})
        return items

    # --- Component Catalogues ---
    def get_catalogue(self, cat_type):
        """Return a component catalogue."""
        cats = {
            'eps': EPS_CATALOGUE,
            'radio': RADIO_CATALOGUE,
            'adcs': ADCS_CATALOGUE,
            'obc': OBC_CATALOGUE,
        }
        return json.dumps({"status": "ok", "items": cats.get(cat_type, []), "type": cat_type})

    def select_component(self, cat_type, model):
        """Select a component from any catalogue."""
        cats = {'eps': EPS_CATALOGUE, 'radio': RADIO_CATALOGUE, 'adcs': ADCS_CATALOGUE, 'obc': OBC_CATALOGUE}
        catalogue = cats.get(cat_type, [])
        for item in catalogue:
            if item['model'] == model:
                if cat_type == 'eps':
                    self.selected_eps = item
                elif cat_type == 'radio':
                    self.selected_radio = item
                elif cat_type == 'adcs':
                    self.selected_adcs = item
                elif cat_type == 'obc':
                    self.selected_obc = item
                self._update_bom(cat_type.upper(), item['vendor'], item['model'],
                                item['mass_g'], item['cost_usd'])
                return json.dumps({"status": "ok", "selected": item})
        return json.dumps({"status": "error"})

    # --- BOM ---
    def _update_bom(self, category, vendor, model, mass_g, cost_usd):
        """Update BOM with selected component."""
        # Remove existing item in same category
        self.bom_items = [b for b in self.bom_items if b['category'] != category]
        self.bom_items.append({
            'category': category, 'vendor': vendor, 'model': model,
            'mass_g': mass_g, 'cost_usd': cost_usd, 'qty': 1
        })

    def get_bom(self):
        """Return full BOM."""
        # Add solar panel cost
        budget = calc_power_budget(self.subsystems, self.orbit)
        solar = calc_solar_panel(budget['solar_required_w'], self.solar_config.get('cell_type', 'Azur 3G30A'),
                                 self.mission_years, self.solar_config.get('margin_pct', 20))

        bom = list(self.bom_items)
        # Check if solar already in BOM
        if not any(b['category'] == 'Solar' for b in bom) and solar['cost_usd'] > 0:
            bom.append({
                'category': 'Solar', 'vendor': 'Calculated',
                'model': f"{solar['cell_type']} ({solar['num_cells']} cells)",
                'mass_g': round(solar['mass_kg'] * 1000), 'cost_usd': solar['cost_usd'], 'qty': 1
            })

        total_cost = sum(b['cost_usd'] * b.get('qty', 1) for b in bom)
        total_mass = sum(b['mass_g'] * b.get('qty', 1) for b in bom)

        return json.dumps({
            "status": "ok",
            "items": bom,
            "total_cost_usd": round(total_cost),
            "total_mass_g": round(total_mass),
            "total_mass_kg": round(total_mass / 1000, 3),
        })

    def add_bom_item(self, category, vendor, model, mass_g, cost_usd, qty):
        """Add custom item to BOM."""
        self.bom_items.append({
            'category': category, 'vendor': vendor, 'model': model,
            'mass_g': _safe_float(mass_g), 'cost_usd': _safe_float(cost_usd),
            'qty': int(_safe_float(qty, 1))
        })
        return json.dumps({"status": "ok"})

    def remove_bom_item(self, index):
        """Remove BOM item."""
        idx = int(index)
        if 0 <= idx < len(self.bom_items):
            self.bom_items.pop(idx)
            return json.dumps({"status": "ok"})
        return json.dumps({"status": "error"})

    # --- Import Power Budget from Excel ---
    def import_power_budget(self):
        """Import power budget from Excel file."""
        file_types = ('Excel Files (*.xlsx;*.xls)', 'CSV Files (*.csv)', 'All Files (*.*)')
        result = self.window.create_file_dialog(webview.OPEN_DIALOG, allow_multiple=False, file_types=file_types)
        if not result or len(result) == 0:
            return json.dumps({"status": "cancelled"})
        filepath = result[0]
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            self.subsystems = []
            # Find header row
            header_row = 1
            col_map = {}
            for row_idx in range(1, min(10, ws.max_row + 1)):
                for cell in ws[row_idx]:
                    val = str(cell.value or '').strip().lower()
                    ci = cell.column - 1
                    if 'name' in val or 'subsystem' in val or 'component' in val:
                        col_map['name'] = ci
                        header_row = row_idx
                    elif 'safe' in val:
                        col_map['safe_w'] = ci
                    elif 'nominal' in val or 'typical' in val:
                        col_map['nominal_w'] = ci
                    elif 'peak' in val or 'max' in val:
                        col_map['peak_w'] = ci
                    elif 'duty' in val:
                        col_map['duty_pct'] = ci
                    elif 'category' in val or 'type' in val:
                        col_map['category'] = ci

            if 'name' not in col_map:
                wb.close()
                return json.dumps({"status": "error", "message": "Could not find subsystem Name column."})

            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
                cells = [c.value for c in row]
                name_idx = col_map['name']
                if name_idx >= len(cells) or not cells[name_idx]:
                    continue
                name = str(cells[name_idx]).strip()
                if not name:
                    continue
                self.subsystems.append({
                    'name': name,
                    'category': str(cells[col_map.get('category', name_idx)] or 'Other').strip() if 'category' in col_map else 'Other',
                    'safe_w': _safe_float(cells[col_map.get('safe_w', -1)] if 'safe_w' in col_map and col_map['safe_w'] < len(cells) else 0),
                    'nominal_w': _safe_float(cells[col_map.get('nominal_w', -1)] if 'nominal_w' in col_map and col_map['nominal_w'] < len(cells) else 0),
                    'peak_w': _safe_float(cells[col_map.get('peak_w', -1)] if 'peak_w' in col_map and col_map['peak_w'] < len(cells) else 0),
                    'duty_pct': _safe_float(cells[col_map.get('duty_pct', -1)] if 'duty_pct' in col_map and col_map['duty_pct'] < len(cells) else 100, 100),
                })
            wb.close()
            return json.dumps({"status": "ok", "count": len(self.subsystems), "filename": os.path.basename(filepath)})
        except Exception as e:
            return json.dumps({"status": "error", "message": str(e)})

    # --- Export ---
    def export_excel(self):
        """Export full mission data to Excel."""
        result = self.window.create_file_dialog(
            webview.SAVE_DIALOG,
            save_filename=f"{self.mission_name.replace(' ', '_')}_BOM.xlsx",
            file_types=('Excel Files (*.xlsx)',)
        )
        if not result:
            return json.dumps({"status": "cancelled"})
        filepath = result if isinstance(result, str) else result[0]
        try:
            wb = openpyxl.Workbook()
            hfill = openpyxl.styles.PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            hfont = openpyxl.styles.Font(bold=True, color="FFFFFF", size=10)
            afill = openpyxl.styles.PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

            def wh(ws, row, headers):
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=c, value=h)
                    cell.font = hfont; cell.fill = hfill; cell.border = border

            def aw(ws):
                for col in ws.columns:
                    try:
                        cl = col[0].column_letter
                    except AttributeError:
                        continue
                    mx = max((len(str(c.value or '')) for c in col if not isinstance(c, openpyxl.cell.cell.MergedCell)), default=8)
                    ws.column_dimensions[cl].width = min(mx + 3, 40)

            # Sheet 1: Mission Summary
            ws = wb.active
            ws.title = "Mission Summary"
            ws['A1'] = f'CubeSat Mission: {self.mission_name}'
            ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            info = [
                ('Form Factor', self.form_factor),
                ('Max Mass (kg)', CUBESAT_STANDARDS.get(self.form_factor, {}).get('mass_kg', '')),
                ('Orbit Altitude (km)', self.orbit.get('altitude_km', '')),
                ('Orbit Period (min)', self.orbit.get('period_min', '')),
                ('Eclipse (min)', self.orbit.get('eclipse_min', '')),
                ('Sunlight (min)', self.orbit.get('sunlight_min', '')),
                ('Mission Duration (years)', self.mission_years),
            ]
            for i, (k, v) in enumerate(info):
                ws.cell(row=3+i, column=1, value=k).border = border
                ws.cell(row=3+i, column=2, value=v).border = border
            aw(ws)

            # Sheet 2: Power Budget
            ws_p = wb.create_sheet("Power Budget")
            ws_p['A1'] = 'Power Budget'
            ws_p['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            wh(ws_p, 3, ['Subsystem', 'Category', 'Safe (W)', 'Nominal (W)', 'Peak (W)', 'Duty (%)', 'Avg (W)'])
            for i, s in enumerate(self.subsystems):
                r = i + 4
                avg = s['nominal_w'] * s['duty_pct'] / 100
                for c, v in enumerate([s['name'], s.get('category', ''), s['safe_w'], s['nominal_w'],
                                       s['peak_w'], s['duty_pct'], round(avg, 2)], 1):
                    cell = ws_p.cell(row=r, column=c, value=v)
                    cell.border = border
                    if i % 2: cell.fill = afill
            budget = calc_power_budget(self.subsystems, self.orbit)
            r = len(self.subsystems) + 5
            for k, v in budget.items():
                if k == 'margin_note': continue
                ws_p.cell(row=r, column=1, value=k.replace('_', ' ').title()).border = border
                ws_p.cell(row=r, column=2, value=v).border = border
                r += 1
            aw(ws_p)

            # Sheet 3: Mass Budget
            ws_m = wb.create_sheet("Mass Budget")
            ws_m['A1'] = 'Mass Budget'
            ws_m['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            mass_items = self._build_mass_list()
            wh(ws_m, 3, ['Component', 'Category', 'Mass (kg)', 'Source'])
            for i, m in enumerate(mass_items):
                r = i + 4
                for c, v in enumerate([m['name'], m['category'], m['mass_kg'], m['source']], 1):
                    cell = ws_m.cell(row=r, column=c, value=v)
                    cell.border = border
                    if i % 2: cell.fill = afill
            total_mass = sum(m['mass_kg'] for m in mass_items)
            max_mass = CUBESAT_STANDARDS.get(self.form_factor, {}).get('mass_kg', 4.0)
            r = len(mass_items) + 5
            ws_m.cell(row=r, column=1, value='Total').font = openpyxl.styles.Font(bold=True)
            ws_m.cell(row=r, column=3, value=round(total_mass, 3)).font = openpyxl.styles.Font(bold=True)
            ws_m.cell(row=r+1, column=1, value='Max Allowed')
            ws_m.cell(row=r+1, column=3, value=max_mass)
            ws_m.cell(row=r+2, column=1, value='Margin')
            ws_m.cell(row=r+2, column=3, value=round(max_mass - total_mass, 3))
            aw(ws_m)

            # Sheet 4: BOM
            ws_b = wb.create_sheet("BOM")
            ws_b['A1'] = 'Bill of Materials'
            ws_b['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            bom_data = json.loads(self.get_bom())
            wh(ws_b, 3, ['Category', 'Vendor', 'Model', 'Qty', 'Mass (g)', 'Unit Cost (USD)', 'Total Cost (USD)'])
            for i, b in enumerate(bom_data['items']):
                r = i + 4
                vals = [b['category'], b['vendor'], b['model'], b.get('qty', 1),
                        b['mass_g'], b['cost_usd'], b['cost_usd'] * b.get('qty', 1)]
                for c, v in enumerate(vals, 1):
                    cell = ws_b.cell(row=r, column=c, value=v)
                    cell.border = border
                    if i % 2: cell.fill = afill
            r = len(bom_data['items']) + 5
            ws_b.cell(row=r, column=1, value='TOTAL').font = openpyxl.styles.Font(bold=True, size=11)
            ws_b.cell(row=r, column=5, value=bom_data['total_mass_g']).font = openpyxl.styles.Font(bold=True)
            ws_b.cell(row=r, column=7, value=bom_data['total_cost_usd']).font = openpyxl.styles.Font(bold=True)
            ws_b.cell(row=r, column=7).number_format = '#,##0'
            aw(ws_b)

            wb.save(filepath)
            wb.close()
            return json.dumps({"status": "ok", "filepath": filepath})
        except Exception as e:
            import traceback
            return json.dumps({"status": "error", "message": str(e), "trace": traceback.format_exc()})


def get_html_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'CubeSat-Builder.html')


if __name__ == '__main__':
    api = Api()
    window = webview.create_window(
        'CubeSat Builder',
        get_html_path(),
        js_api=api,
        width=1400,
        height=900,
        min_size=(1000, 600)
    )
    api.window = window
    webview.start(debug=False)
